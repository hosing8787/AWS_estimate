"""
견적서 생성 스크립트
원본 완전 분석 기반 - 테두리/수식/병합/색상 정확 재현
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, Color
from openpyxl.styles.fills import PatternFill
import os
import urllib.parse
import urllib.request
import json
import zipfile

def fetch_azure_price(region, service_name, meter_name, sku_name, product_name):
    filters = [
        f"armRegionName eq '{region}'",
        f"serviceName eq '{service_name}'",
        f"priceType eq 'Consumption'"
    ]
    if meter_name: filters.append(f"meterName eq '{meter_name}'")
    if sku_name: filters.append(f"skuName eq '{sku_name}'")
    if product_name: filters.append(f"productName eq '{product_name}'")
    
    filter_str = ' and '.join(filters)
    url = f"https://prices.azure.com/api/retail/prices?$filter={urllib.parse.quote(filter_str)}"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode())
            items = data.get('Items', [])
            if items:
                return float(items[0].get('retailPrice', 0))
    except Exception as e:
        print(f"Error fetching price: {e}")
    return 0.0

# ── 공통 유틸 ──────────────────────────────────────────────────────────────────

def S(style=None):
    return Side(style=style)

def brd(l=None, r=None, t=None, b=None):
    return Border(left=S(l), right=S(r), top=S(t), bottom=S(b))

def rgb_fill(hex6):
    return PatternFill('solid', fgColor=f'FF{hex6}')

def theme_fill(n=0, tint=0.0):
    c = Color(theme=n, tint=tint, type='theme')
    pf = PatternFill(patternType='solid')
    pf.fgColor = c
    return pf

TF0_DARK = None

def tf0():
    return theme_fill(0, 0.0)

def tf0d():
    return theme_fill(0, -0.1499984740745262)

def font(bold=False, size=10, name='맑은 고딕'):
    return Font(name=name, bold=bold, size=size)

def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def port_styles_from_original(output_path, orig_path):
    import zipfile, os
    from xml.etree import ElementTree as ET

    NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'

    with zipfile.ZipFile(orig_path) as z:
        orig_styles = z.read('xl/styles.xml')

        def extract_maps(xml_bytes):
            s_map = {}
            v_map = {}
            root = ET.fromstring(xml_bytes)
            for row_elem in root.findall(f'.//{NS}row'):
                for cell in row_elem.findall(f'{NS}c'):
                    ref = cell.get('r')
                    s = cell.get('s')
                    if s:
                        s_map[ref] = s
                    f_elem = cell.find(f'{NS}f')
                    v_elem = cell.find(f'{NS}v')
                    if f_elem is not None and v_elem is not None and v_elem.text is not None:
                        v_map[ref] = (v_elem.text, cell.get('t', ''))
            return s_map, v_map

        orig_s1, orig_v1 = extract_maps(z.read('xl/worksheets/sheet1.xml'))
        orig_s2, orig_v2 = extract_maps(z.read('xl/worksheets/sheet2.xml'))

    with zipfile.ZipFile(output_path) as z:
        files = {name: z.read(name) for name in z.namelist()}

    files['xl/styles.xml'] = orig_styles

def port_styles_from_original(wb_gen, orig_path):
    from openpyxl import load_workbook
    from copy import copy
    wb_orig = load_workbook(orig_path)

    # Port Sheet 1
    ws_gen1 = wb_gen['매출견적서']
    ws_orig1 = wb_orig['매출견적서']
    for r in range(1, ws_gen1.max_row + 1):
        for c in range(1, 13):
            orig_c = ws_orig1.cell(row=r, column=c)
            gen_c = ws_gen1.cell(row=r, column=c)
            if orig_c.has_style:
                gen_c.font = copy(orig_c.font)
                gen_c.border = copy(orig_c.border)
                gen_c.fill = copy(orig_c.fill)
                gen_c.number_format = copy(orig_c.number_format)
                gen_c.alignment = copy(orig_c.alignment)



# ==============================================================================
# 메인
# ==============================================================================

def sc(ws, coord, val=None, fnt=None, al=None, border=None, fill=None, nf=None):
    c = ws[coord]
    if val is not None:
        c.value = val
    if fnt:
        c.font = fnt
    if al:
        c.alignment = al
    if border is not None:
        c.border = border
    if fill is not None:
        c.fill = fill
    if nf:
        c.number_format = nf
    return c

WON_FMT  = r'_-"₩"* #,##0_-;\\-"₩"* #,##0_-;_-"₩"* "-"_-;_-@_-'
WON_FMT2 = r'"₩"#,##0_);[Red]\("₩"#,##0\)'
USD_FMT  = r'\$#,##0.00_);[Red]\(\$#,##0.00\)'
QTY_FMT  = r'0_);[Red]\(0\)'

T4 = brd('thin','thin','thin','thin')
H4 = brd('hair','hair','hair','hair')
M4 = brd('medium','medium','medium','medium')

# ==============================================================================
# 시트 1 : 매출견적서
# ==============================================================================

def build_sheet1(wb, subtotals):
    ws = wb.active
    ws.title = '매출견적서'

    for col, w in [('A',1.375),('B',11.625),('C',32.625),('D',47.5),
                   ('E',6.125),('F',15.625),('G',13.0),('H',53.375),
                   ('I',1.375),('J',9.0)]:
        ws.column_dimensions[col].width = w

    for r, h in [(2,33.75),(3,33.75),(6,18),(7,18),(8,18),(9,18),(10,18),(11,18),
                 (14,19.5),(15,30.75)]:
        ws.row_dimensions[r].height = h

    ws.merge_cells('B2:H2')
    sc(ws,'B2','見   積   書',
       fnt=Font(name='맑은 고딕',bold=True,size=26),
       al=aln('center','center'))

    # Will be populated by port_styles_from_original
    # But just in case, we set it up structurally
    for r, txt in [(6,'회   사   명 : '),(7,'견 적 번 호 : 가견적'),
                   (8,'견 적 날 짜 : 2026년  1월  6일'),
                   (9,'견적 유효 기간 : 견적일로부터 30일'),
                   (10,'납         기 : 별도 협의')]:
        sc(ws,f'B{r}',txt,fnt=font(),al=aln('left','center'))

    ws.merge_cells('E6:E11')
    sc(ws,'E6','공\n\n급\n\n자',fnt=font(),al=aln('center','center',wrap=True),
       border=brd('dotted','dotted','dotted','dotted'), fill=theme_fill(0))

    # E column has merged cells so only the bottom of E11 needs a border, but sc only targets E6
    # Let's apply proper dotted borders to the entire block E6:H11
    for r in range(6, 12):
        # E column borders (merged, so we just set left/right/top/bottom for the whole merged block if we could, but let's apply to each cell)
        sc(ws, f'E{r}', border=brd('dotted', 'dotted', 'dotted' if r==6 else None, 'dotted' if r==11 else None))
        
        # Merge G and H
        ws.merge_cells(f'G{r}:H{r}')
        
        # F column (labels)
        sc(ws, f'F{r}', border=brd('dotted', 'dotted', 'dotted', 'dotted'), fill=theme_fill(0))
        
        # G column (values) - applies to G:H
        sc(ws, f'G{r}', border=brd('dotted', 'dotted', 'dotted', 'dotted'), fill=theme_fill(0))
        sc(ws, f'H{r}', border=brd('dotted', 'dotted', 'dotted', 'dotted'), fill=theme_fill(0))

    for coord, txt in [('F6','등록번호'),('F7','상       호'),('F8','대표자명'),
                       ('F9','주       소'),('F10','업       태'),('F11','종       목')]:
        sc(ws,coord,txt,fnt=font(),al=aln('left','center'))
        
    for coord, txt in [('G6','  123-45-67890'),('G7','  SK㈜'),
                       ('G8','  대표이사 (직인생략)'),
                       ('G9','  서울특별시 강남구 테헤란로 123'),
                       ('G10','  서비스(사업관련)'),
                       ('G11','  기타정보처리및컴퓨터 운용관련')]:
        sc(ws,coord,txt,fnt=font(),al=aln('left','center'))

    sc(ws,'B14','제   목 : ',fnt=font(),al=aln('left','center'))
    sc(ws,'C14','클라우드 인프라 환경 구축',fnt=font(),al=aln('left','center'))
    sc(ws,'H14','(금액단위 : 원/VAT 별도)',fnt=font(),al=aln('right','center'))

    hf = rgb_fill('F99107')
    ws.merge_cells('C15:D15')
    for coord, txt in [('B15','No.'),('E15','수량\n(개월)'),('F15','월 공급금액'),
                       ('G15','총 공급금액\n(1년 간)'),('H15','비고')]:
        sc(ws,coord,txt,fnt=font(bold=True),al=aln('center','center',wrap=True),border=T4,fill=hf)
    sc(ws,'C15','품목 및 사양',fnt=font(bold=True),al=aln('center','center'),border=T4,fill=hf)
    sc(ws,'D15',border=brd('thin','thin','thin','thin'))

    current_az_row = 16
    rows_az = []
    
    for i, (main_cat, sub_row) in enumerate(subtotals):
        desc = None
        if i == 0:
            desc = '-Azure Resource List Price의 10% 할인 적용\n(Pay-as-you-go 상품에 한하며, RI상품은 5%할인 적용됨)'
        rows_az.append((
            current_az_row, main_cat, 'Cloud 사용료', 12, 
            f"='Azure Resource'!K{sub_row}", f'=F{current_az_row}*E{current_az_row}', desc
        ))
        current_az_row += 1
        
    asp_row = current_az_row
    sum_range = f'F16:F{asp_row-1}' if asp_row > 16 else '0'
    rows_az.append((
        asp_row, 'Azure Support Plan', 'Enterprise Support for Azure',
        12, f'=SUM({sum_range})*0.095', f'=F{asp_row}*E{asp_row}', None
    ))
    
    ws.merge_cells(f'B16:B{asp_row}')
    sc(ws,'B16','Azure\nResource',fnt=font(bold=True),
       al=aln('center','center',wrap=True),border=T4,fill=tf0d())
    for r in range(17, asp_row):
        sc(ws,f'B{r}',border=brd('thin','thin',None,None))
    sc(ws,f'B{asp_row}',border=brd('thin','thin',None,'thin'))
    
    for row,c,d,e,f,g,h in rows_az:
        ws.row_dimensions[row].height = 30.75
        sc(ws,f'C{row}',c,fnt=font(),al=aln('center','center'),border=T4)
        sc(ws,f'D{row}',d,fnt=font(),al=aln('center','center'),border=T4)
        sc(ws,f'E{row}',e,fnt=font(),al=aln('center','center'),border=T4,nf=QTY_FMT)
        sc(ws,f'F{row}',f,fnt=font(),al=aln('center','center'),border=T4,nf=WON_FMT)
        sc(ws,f'G{row}',g,fnt=font(),al=aln('center','center'),border=T4,nf=WON_FMT)
        
        if h:
            hs, he = row, row+1
            if row == asp_row: hs, he = row, row 
            if hs < he:
                ws.merge_cells(f'H{hs}:H{he}')
                sc(ws,f'H{hs}',h,fnt=font(),al=aln('left','center',wrap=True),
                   border=brd('thin','thin','thin',None))
                sc(ws,f'H{he}',border=brd('thin','thin',None,'thin'))
            else:
                sc(ws,f'H{row}',h,fnt=font(),al=aln('left','center',wrap=True),
                   border=brd('thin','thin','thin','thin'))
        else:
            if ws[f'H{row}'].value is None and not type(ws[f'H{row}']).__name__ == 'MergedCell':
                sc(ws,f'H{row}',border=brd('thin','thin',None if row>16 else 'thin',None if row<asp_row else 'thin'))

    az_total_row = asp_row + 1
    ws.row_dimensions[az_total_row].height = 30.75
    ws.merge_cells(f'B{az_total_row}:E{az_total_row}')
    sc(ws,f'B{az_total_row}','Azure Resource 합계',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=tf0d())
    for col in ['C','D','E']: sc(ws,f'{col}{az_total_row}',border=brd('thin','thin','thin','thin'))
    sc(ws,f'F{az_total_row}',f'=ROUNDDOWN(SUM(F16:F{asp_row}),-3)',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=tf0d(),nf=WON_FMT)
    sc(ws,f'G{az_total_row}',f'=ROUNDDOWN(SUM(G16:G{asp_row}),-3)',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=tf0d(),nf=WON_FMT)
    sc(ws,f'H{az_total_row}',border=T4,fill=tf0d())
    
    msp_start = az_total_row + 1
    msp_end = msp_start + 2
    ws.merge_cells(f'B{msp_start}:B{msp_end}')
    sc(ws,f'B{msp_start}','MSP\n운영비',fnt=font(bold=True),
       al=aln('center','center',wrap=True),border=T4,fill=tf0d())
    sc(ws,f'B{msp_start+1}',border=brd('thin','thin',None,None))
    sc(ws,f'B{msp_end}',border=brd('thin','thin',None,'thin'))
    
    ws.row_dimensions[msp_start].height = 30.75
    ws.row_dimensions[msp_start+1].height = 30.75
    ws.row_dimensions[msp_end].height = 24

    rows_msp = [
        (msp_start,'IaaS Management (Premium)','Cloud 사용료의 21%',12,f'=SUM({sum_range})*0.21',f'=F{msp_start}*E{msp_start}'),
        (msp_start+1,'Enterprise IT Management', 'Cloud 사용료의 7%', 12,f'=SUM({sum_range})*0.07',f'=F{msp_start+1}*E{msp_start+1}'),
        (msp_end,'Expert Infra',             'DB/MW/3rd Party 운영비용',12,'=(170000*0)+(40000*0)',f'=F{msp_end}*E{msp_end}'),
    ]
    for row,c,d,e,f,g in rows_msp:
        sc(ws,f'C{row}',c,fnt=font(),al=aln('center','center'),border=T4,fill=theme_fill(0))
        sc(ws,f'D{row}',d,fnt=font(),al=aln('center','center'),border=T4,fill=theme_fill(0))
        sc(ws,f'E{row}',e,fnt=font(),al=aln('center','center'),border=T4,nf=QTY_FMT)
        sc(ws,f'F{row}',f,fnt=font(),al=aln('center','center'),border=T4,nf=WON_FMT)
        sc(ws,f'G{row}',g,fnt=font(),al=aln('center','center'),border=T4,nf=WON_FMT)

    sc(ws,f'H{msp_start}',border=T4)
    sc(ws,f'H{msp_start+1}',border=brd('thin','thin',None,'thin'))
    sc(ws,f'H{msp_end}','- DB Expert Infra 수량 0EA (개발0/운영0)\n',
       fnt=font(),al=aln('left','center',wrap=True),
       border=brd('thin','thin','thin',None))

    msp_tot_row = msp_end + 1
    ws.row_dimensions[msp_tot_row].height = 30.75
    ws.merge_cells(f'B{msp_tot_row}:E{msp_tot_row}')
    sc(ws,f'B{msp_tot_row}','MSP 운영비 합계',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=tf0d())
    for col in ['C','D','E']: sc(ws,f'{col}{msp_tot_row}',border=brd('thin','thin','thin','thin'))
    sc(ws,f'F{msp_tot_row}',f'=ROUNDDOWN(SUM(F{msp_start}:F{msp_end}),-3)',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=tf0d(),nf=WON_FMT)
    sc(ws,f'G{msp_tot_row}',f'=ROUNDDOWN(SUM(G{msp_start}:G{msp_end}),-3)',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=tf0d(),nf=WON_FMT)
    sc(ws,f'H{msp_tot_row}',border=T4,fill=tf0d())

    blank_row = msp_tot_row + 1
    ws.row_dimensions[blank_row].height = 24
    ws.merge_cells(f'B{blank_row}:H{blank_row}')
    sc(ws,f'B{blank_row}','이하 여백',fnt=font(),al=aln('center','center'),
       border=brd(None,'thin','thin','thin'),fill=theme_fill(0))
    for col in ['C','D','E','F','G']: sc(ws,f'{col}{blank_row}',border=brd('thin','thin','thin','thin'))
    sc(ws,f'H{blank_row}',border=brd('thin',None,'thin','thin'))

    tot1_row = blank_row + 1
    ws.row_dimensions[tot1_row].height = 23.25
    ws.merge_cells(f'B{tot1_row}:E{tot1_row}')
    yf = rgb_fill('FFFF99')
    sc(ws,f'B{tot1_row}','사용료 합계(부가세 별도)',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=yf)
    for col in ['C','D','E']: sc(ws,f'{col}{tot1_row}',border=brd('thin','thin','thin','thin'),fill=yf)
    sc(ws,f'F{tot1_row}',f'=SUM(F{az_total_row},F{msp_tot_row})',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=yf,nf=WON_FMT)
    sc(ws,f'G{tot1_row}',f'=SUM(G{az_total_row},G{msp_tot_row})',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=yf,nf=WON_FMT)
    sc(ws,f'H{tot1_row}',border=T4,fill=yf)

    tot2_row = tot1_row + 1
    ws.row_dimensions[tot2_row].height = 23.25
    ws.merge_cells(f'B{tot2_row}:E{tot2_row}')
    of = rgb_fill('FFCC99')
    sc(ws,f'B{tot2_row}','최종 사용료 - 최종 계약금(VAT 별도)',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=of)
    for col in ['C','D','E']: sc(ws,f'{col}{tot2_row}',border=brd('thin','thin','thin','thin'),fill=of)
    sc(ws,f'F{tot2_row}',f'=F{tot1_row}',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=of,nf=WON_FMT)
    sc(ws,f'G{tot2_row}',f'=G{tot1_row}',fnt=font(bold=True),
       al=aln('center','center'),border=T4,fill=of,nf=WON_FMT)
    sc(ws,f'H{tot2_row}',border=T4,fill=of)

    rmk_row1 = tot2_row + 1
    rmk_row2 = rmk_row1 + 1
    ws.row_dimensions[rmk_row1].height = 21
    ws.row_dimensions[rmk_row2].height = 15
    sc(ws,f'B{rmk_row1}','※ Remark',fnt=font(bold=True))
    sc(ws,f'B{rmk_row2}','- 상기 비용은 실 사용량 및 환율 변동에 따라 금액이 변동될 수 있음.',fnt=font())


# ==============================================================================
# 시트 2 : Azure Resource
# ==============================================================================

def build_sheet2(wb):
    ws = wb.create_sheet('Azure Resource')

    for col, w in [('A',1.625),('B',11.625),('C',32.875),('D',24.125),('E',26.375),
                   ('F',44.25),('G',6.25),('H',5.875),('I',12.875),('J',13.0),
                   ('K',15.125),('L',59.125),('M',1.625),('N',9.0)]:
        ws.column_dimensions[col].width = w

    for r, h in [(2,26.25),(3,14.25),(4,14.25),(5,16.5),(6,17.25)]:
        ws.row_dimensions[r].height = h

    sc(ws,'B2','Cloud 세부내역 : ',
       fnt=font(bold=True,size=12),al=aln('left','center'))
    sc(ws,'K3','환율',fnt=font(bold=True),al=aln('center','center'),border=T4)
    for col in ['C','D','E','F']:
        sc(ws,f'{col}4',border=brd(None,None,None,'medium'))
    sc(ws,'J4',border=brd(None,'thin',None,'medium'))
    sc(ws,'K4',1430,fnt=font(),al=aln('center','center'),border=brd(None,'thin',None,'medium'))

    for rng in ['B5:B6','C5:F5','G5:G6','H5:H6','I5:I6','J5:J6','K5:K6','L5:L6']:
        ws.merge_cells(rng)
    ws.merge_cells('D6:F6')

    hf = rgb_fill('F99107')

    sc(ws,'B5','유형',fnt=font(bold=True),al=aln('center','center',wrap=True),
       border=brd('medium','thin','medium',None),fill=hf)
    sc(ws,'C5','품목 및 사양',fnt=font(bold=True),al=aln('center','center'),
       border=brd('thin',None,'medium',None),fill=hf)
    for col in ['D','E','F']:
        sc(ws,f'{col}5',border=brd(None,None,'medium',None))
    sc(ws,'G5','수량',fnt=font(bold=True),al=aln('center','center',wrap=True),
       border=brd('thin',None,'medium',None),fill=hf)
    for coord, txt in [('H5','단위'),('I5','Price\n($/Month)'),
                       ('J5','Price\n(\\/Month)'),('K5','C&C 공급금액\n(\\/Month)')]:
        sc(ws,coord,txt,fnt=font(bold=True),al=aln('center','center',wrap=True),
           border=brd('thin','thin','medium',None),fill=hf)
    sc(ws,'L5','비고',fnt=font(bold=True),al=aln('center','center',wrap=True),
       border=brd(None,'medium','medium',None),fill=theme_fill(9))

    sc(ws,'B6',border=brd('medium','thin',None,'medium'))
    sc(ws,'C6','서비스 명',fnt=font(bold=True),al=aln('center','center'),
       border=brd('thin',None,'thin','medium'),fill=hf)
    sc(ws,'D6','상세 Spec',fnt=font(bold=True),al=aln('center','center'),
       border=brd(None,None,'thin','medium'),fill=hf)
    sc(ws,'E6',border=brd(None,None,'thin','medium'))
    sc(ws,'F6',border=brd(None,'thin','thin','medium'))
    sc(ws,'G6',border=brd('thin',None,None,'medium'))
    for col in ['H','I','J','K']:
        sc(ws,f'{col}6',border=brd('thin','thin',None,'medium'))
    sc(ws,'L6',border=brd(None,'medium',None,'medium'))

    try:
        with open('sample_data.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            inventory = data.get('inventory_items', [])
    except Exception as e:
        print(f"Failed to load sample_data.json: {e}")
        inventory = []

    # --- Dynamic Hierarchy Engine ---
    final_hierarchy = {}
    for item in inventory:
        g_path = item.get('group_path', '기타자원 > 기타')
        parts = [p.strip() for p in g_path.split('>')]
        main_cat = parts[0]
        sub_cat = parts[1] if len(parts) > 1 else '기타'
        
        if main_cat not in final_hierarchy:
            final_hierarchy[main_cat] = {}
        if sub_cat not in final_hierarchy[main_cat]:
            final_hierarchy[main_cat][sub_cat] = []
        final_hierarchy[main_cat][sub_cat].append(item)

    def get_brd(code):
        m = {'tm': brd('thin','thin','medium','thin'),
             'tm_nobot': brd('thin','thin','medium',None),
             'tt': brd('thin','thin','thin','thin'),
             'th_E': brd('thin','hair','thin','thin'),
             'th_F': brd('hair','thin','thin','thin'),
             'tt_hairB': brd('thin','thin','thin','hair'),
             'tt_noB': brd('thin','thin','thin',None),
             'tt_noT': brd('thin','thin',None,'thin'),
             'tt_noTB': brd('thin','thin',None,None),
             'K_mid': brd('thin',None,None,'thin'),
             'K_lb': brd('thin','thin','thin','medium'),
             'K_last': brd('thin',None,None,'thin'),
             'tm_bot': brd('thin','thin','thin','medium'),
             'tm_noR': brd('thin',None,'thin','medium'),
             'hair_L': brd('hair','thin','thin',None),
             'med_bot': brd('thin','thin','thin','medium'),
             'L_first': brd('thin','medium','medium','thin'),
             'L_mid': brd('thin','medium','thin','thin'),
             'L_lb': brd('thin','medium','thin',None),
             'L_med': brd('thin','medium','thin','thin'),
             'L_last': brd('thin','medium','thin',None),
             'tt_medR': brd('thin','medium','medium','thin'),
             'tt_medR_noB': brd('thin','medium','thin',None),
             'tm_bot_last': brd('thin','thin','thin','medium'),
             'tm_noR_last': brd('thin',None,'thin','medium'),
             'hair_L_last': brd('hair','thin','thin',None),
             'med_bot_last': brd('thin',None,None,'thin'),
             'tt_medR_noB_last': brd('thin','medium','thin',None),
             }
        return m.get(code, T4)

    current_row = 7
    subtotals = []
    
    for main_cat, sub_dict in final_hierarchy.items():
        main_start = current_row
        main_end = current_row + sum(len(lst) for lst in sub_dict.values()) - 1
        
        for sub_cat, items_list in sub_dict.items():
            sub_start = current_row
            sub_end = current_row + len(items_list) - 1
            
            # Merging and Styling C column (Sub-category label)
            if sub_start < sub_end:
                ws.merge_cells(f'C{sub_start}:C{sub_end}')
                sc(ws, f'C{sub_start}', sub_cat, fnt=font(bold=True), al=aln('center','center'),
                   border=brd('thin','thin', 'medium' if sub_start == main_start else 'thin', None))
                for r in range(sub_start+1, sub_end+1):
                    sc(ws, f'C{r}', border=brd('thin','thin',None,'medium' if r==main_end else ('thin' if r==sub_end else None)))
            else:
                sc(ws, f'C{sub_start}', sub_cat, fnt=font(bold=True), al=aln('center','center'),
                   border=brd('thin','thin', 'medium' if sub_start == main_start else 'thin', 'medium' if sub_end == main_end else 'thin'))

            for idx, item in enumerate(items_list):
                r = sub_start + idx
                
                t_style = 'medium' if r == main_start else 'thin'
                b_style = 'medium' if r == main_end else 'thin'
                
                price_usd_month = item.get('retailPrice', 0)
                if price_usd_month == 0:
                    hourly_price = fetch_azure_price(
                        item.get('region'), 
                        item.get('service_name'), 
                        item.get('meterName'), 
                        item.get('skuName'), 
                        item.get('productName')
                    )
                    if item.get('billing_option') == 'Pay-as-you-go' and item.get('service_name') != 'Storage':
                        price_usd_month = hourly_price * 730
                    else: 
                        price_usd_month = hourly_price
                
                d_val = item.get('service_name')
                e_val = item.get('applied')
                
                spec_str = str(item.get('spec') or '').strip()
                note_str = str(item.get('note') or '').strip()
                if note_str:
                    f_val = f"{spec_str}\n{note_str}"
                else:
                    f_val = spec_str
                    
                g_val = item.get('qty', 1)
                h_val = item.get('unit', 'ea')
                
                # Calculate dynamic row height based on text length
                f_lines = 0
                for line in f_val.split('\n'):
                    f_lines += max(1, len(line) // 40 + 1) # F col width ~44
                
                d_lines = 0
                for line in str(d_val or '').split('\n'):
                    d_lines += max(1, len(line) // 22 + 1) # D col width ~24
                    
                total_lines = max(f_lines, d_lines, 1)
                ws.row_dimensions[r].height = max(20, total_lines * 14.5)

                
                sc(ws,f'D{r}',d_val,fnt=font(),al=aln('center','center'),
                   border=brd('thin','thin', t_style, b_style),fill=theme_fill(0))

                sc(ws,f'E{r}',e_val,fnt=font(),al=aln('center','center'),
                   border=brd('thin','hair', t_style, b_style),fill=theme_fill(0))
                sc(ws,f'F{r}',f_val,fnt=font(),al=aln('left','center',wrap=True),
                   border=brd('hair','thin', t_style, b_style),fill=theme_fill(0))
                
                sc(ws,f'G{r}',g_val,fnt=font(),al=aln('center','center'),
                   border=brd('thin','thin', t_style, b_style))
                sc(ws,f'H{r}',h_val,fnt=font(),al=aln('center','center'),
                   border=brd('thin','thin', t_style, b_style))
                sc(ws,f'I{r}',price_usd_month,fnt=font(),al=aln('center','center'),
                   border=brd('thin','thin', t_style, b_style),nf=USD_FMT)
                sc(ws,f'J{r}',f'=I{r}*$K$4*G{r}',
                   fnt=font(),al=aln('center','center'),
                   border=brd('thin','thin', t_style, b_style),nf=WON_FMT2)
                sc(ws,f'K{r}',f'=J{r}*(1-10%)',
                   fnt=font(),al=aln('center','center'),
                   border=brd('thin','thin', t_style, b_style),nf=WON_FMT2)

                # Column L merge check
                sc(ws,f'L{r}',border=brd(None,'medium', t_style, b_style))
                current_row += 1
        
        if main_end > main_start:
            ws.merge_cells(f'B{main_start}:B{main_end}')
            sc(ws,f'B{main_start}',main_cat,fnt=font(bold=True),
               al=aln('center','center',wrap=True),
               border=brd('medium','thin','medium',None))
            for r in range(main_start+1, main_end+1):
                sc(ws,f'B{r}',border=brd('medium','thin',None,'medium' if r==main_end else None))
        else:
            sc(ws,f'B{main_start}',main_cat,fnt=font(bold=True),
               al=aln('center','center',wrap=True),
               border=brd('medium','thin','medium','medium'))

        subtotal_row = current_row
        ws.row_dimensions[current_row].height = 26.25
        ws.merge_cells(f'B{current_row}:J{current_row}')
        sc(ws,f'B{current_row}',f'{main_cat} 합계 :',fnt=font(bold=True),
           al=aln('right','center'),
           border=brd('medium',None,'medium','medium'),fill=theme_fill(5, 0.5999938962981048))
        for col in ['C','D','E','F','G','H','I','J']:
            sc(ws,f'{col}{current_row}',border=brd(None,None,'medium','medium'))
        sc(ws,f'K{current_row}',f'=SUM(K{main_start}:K{current_row-1})',fnt=font(bold=True),
           al=aln('center','center'),
           border=brd(None,None,'medium','medium'),fill=theme_fill(5, 0.5999938962981048),nf=WON_FMT2)
        sc(ws,f'L{current_row}',border=brd(None,'medium','medium','medium'),fill=theme_fill(5, 0.5999938962981048))
        
        subtotals.append((main_cat, subtotal_row))
        current_row += 1

    # Total Row coordinate depends on fixed padding
    ws.row_dimensions[current_row].height = 26.25
    ws.merge_cells(f'B{current_row}:J{current_row}')
    sc(ws,f'B{current_row}','총 합계(백신제외) :',fnt=font(bold=True),
       al=aln('right','center'),
       border=brd('medium',None,'medium','medium'),fill=rgb_fill('FFC000'))
    for col in ['C','D','E','F','G','H','I','J']:
        sc(ws,f'{col}{current_row}',border=brd(None,None,'medium','medium'))
    
    sub_terms = ",".join([f"K{r}" for _, r in subtotals])
    sum_formula = f"=SUM({sub_terms})" if sub_terms else "=0"
    sc(ws,f'K{current_row}',sum_formula,fnt=font(bold=True),
       al=aln('center','center'),
       border=brd(None,None,'medium','medium'),fill=rgb_fill('FFC000'),nf=WON_FMT2)
    sc(ws,f'L{current_row}',border=brd(None,'medium','medium','medium'),fill=rgb_fill('FFC000'))

    current_row += 2
    # Ensure Remark block starts at row 28 if data is standard
    # In the template, Remark starts at B28. 
    # With 12 + 4 + 1 items and 3 subtotal rows and 1 total row starting from row 7:
    # Row 7-18: Server
    # Row 19: Server Subtotal
    # Row 20-23: 기타자원
    # Row 24: 기타자원 Subtotal
    # Row 25: 모니터링
    # Row 26: 모니터링 Subtotal
    # Row 27: 총합계
    # Row 28: Empty
    # Row 29: Remark starts.
    # Wait, original template has Remark at B28. Let me check my check_rows.py output.
    # Template: Row 26 Total, Row 27 Empty? No, check_rows said B27 is Total Sum. B28 is blank.
    # Actually B28 in template is blank. B29 might be Remark.
    for r in range(current_row, current_row+3):
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f'B{r}:L{r}')
        for col in ['B','C','D','E','F','G','H','I','J','K','L']:
            sc(ws, f'{col}{r}', border=brd(None, None, None, None))
            
    sc(ws,f'B{current_row}','※ Remark',fnt=font(bold=True))
    for col in ['B','C','D','E','F','G','H','I','J','K','L']:
        ws[f'{col}{current_row}'].alignment = aln('left','center')

    sc(ws,f'B{current_row+1}','- 상기 금액은 환율 1430.00원 기준임.',fnt=font())
    for col in ['B','C','D','E','F','G','H','I','J','K','L']:
        ws[f'{col}{current_row+1}'].alignment = aln('left','center')

    sc(ws,f'B{current_row+2}',
       '- 본 견적 금액은 상기 서비스 항목에 대한 자원 공급 금액이며(자원 생성/관리 등의 Managed Service 비용은 제외), 상기 내역 기준에서 기술되지 않은 서비스/기능 및 자원 Spec./수량 등의 변경 시(추가/삭제 등) 견적 금액 재산정이 필요함.',fnt=font())
    for col in ['B','C','D','E','F','G','H','I','J','K','L']:
        ws[f'{col}{current_row+2}'].alignment = aln('left','top',wrap=True)

    ws.row_dimensions[current_row+2].height = 60

    return subtotals

# ==============================================================================
# 메인
# ==============================================================================

def run_generation(output='가상의 고객 주식회사_클라우드 인프라 마이그레이션_최종.xlsx', orig='template.xlsx'):
    wb = Workbook()
    subtotals = build_sheet2(wb)
    build_sheet1(wb, subtotals)

    if os.path.exists(orig):
        port_styles_from_original(wb, orig)
        
    wb.save(output)
    
    # Post-process: inject drawing and media files manually since openpyxl dropped 'em
    if os.path.exists(orig):
        try:
            import xml.etree.ElementTree as ET
            
            with zipfile.ZipFile(orig, 'r') as zin:
                orig_files = {n: zin.read(n) for n in zin.namelist()}
                
            with zipfile.ZipFile(output, 'r') as zout:
                gen_files = {n: zout.read(n) for n in zout.namelist()}
                
            # 1. Copy media & drawings
            for n in orig_files:
                if n.startswith('xl/media/') or n.startswith('xl/drawings/'):
                    gen_files[n] = orig_files[n]
                    
            # 2. Add drawing to [Content_Types].xml
            ct_xml = gen_files['[Content_Types].xml'].decode('utf-8')
            if 'image/jpeg' not in ct_xml:
                ct_xml = ct_xml.replace('<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">', '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="jpeg" ContentType="image/jpeg"/>')
            if '/xl/drawings/drawing1.xml' not in ct_xml:
                ct_xml = ct_xml.replace('</Types>', '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml" /></Types>')
            gen_files['[Content_Types].xml'] = ct_xml.encode('utf-8')
            
            # 3. Add drawing relationship to sheet1
            s1_rels_path = 'xl/worksheets/_rels/sheet1.xml.rels'
            drawing_rid = 'rId1'
            if s1_rels_path in orig_files:
                gen_files[s1_rels_path] = orig_files[s1_rels_path]
                # parse rels to find the rId for drawing1.xml
                rels_xml = orig_files[s1_rels_path].decode('utf-8')
                root = ET.fromstring(rels_xml)
                ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                for rel in root.findall('rels:Relationship', ns):
                    if 'drawing1.xml' in rel.get('Target', ''):
                        drawing_rid = rel.get('Id')
                        break
            else:
                s1_rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>'
                gen_files[s1_rels_path] = s1_rels.encode('utf-8')
                drawing_rid = 'rId1'
                
            # 4. Link drawing in sheet1.xml
            s1_xml = gen_files['xl/worksheets/sheet1.xml'].decode('utf-8')
            
            # Ensure the 'r' namespace is present for r:id
            if 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"' not in s1_xml:
                s1_xml = s1_xml.replace('<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"', '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"')
                
            if f'<drawing r:id="{drawing_rid}"/>' not in s1_xml:
                # Remove any existing drawing tag just in case
                import re
                s1_xml = re.sub(r'<drawing r:id="[^"]+"/>', '', s1_xml)
                s1_xml = s1_xml.replace('</worksheet>', f'<drawing r:id="{drawing_rid}"/></worksheet>')
                gen_files['xl/worksheets/sheet1.xml'] = s1_xml.encode('utf-8')

            # Re-write the output
            with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zout:
                for n, data in gen_files.items():
                    zout.writestr(n, data)

        except Exception as e:
            print(f"Post-processing image inject failed: {e}")
            
    print(f'저장 완료: {output}')

if __name__ == '__main__':
    run_generation()
