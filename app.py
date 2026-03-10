import streamlit as st
import os
import json
from openpyxl import load_workbook
import generate_quote

st.set_page_config(page_title="Azure Quote Converter", layout="centered")

st.title("☁️ Azure Estimate to SSG Quote Template")
st.markdown("Azure 계산기에서 내보낸 엑셀(Exported Estimate.xlsx)을 업로드하면, 지정된 단일 탭 형식의 견적서 파일로 즉시 자동 변환합니다.")

def parse_azure_export(uploaded_file):
    # Load file into openpyxl
    wb = load_workbook(uploaded_file, data_only=True)
    # 엑셀의 첫 번째 시트 사용
    sheet = wb.worksheets[0]
    
    inventory = []
    # 데이터는 보통 4행부터 시작 (1, 2행 제목, 3행 헤더)
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row[0]: # Service category null 이면 스킵
            if len(row) > 3 and row[3] == 'Total':
                break
            continue
            
        cat, svc_type, custom_name, region, desc, monthly_cost, upfront = row[:7]
        
        main_group = "기타자원"
        if cat in ["컴퓨팅", "Virtual Machines", "데이터베이스", "Databases", "Containers"]: 
            main_group = "Server 및 DBMS"
        elif cat in ["Storage", "저장소"]: 
            main_group = "Storage"
        elif cat in ["Networking", "네트워크"]: 
            main_group = "Network"
        elif cat in ["Management and Governance"]:
            main_group = "모니터링"
            
        group_path = f"{main_group} > {cat}"
        
        qty = 1
        unit = "ea"
        if isinstance(desc, str) and " x " in desc:
            try:
                parts = desc.split(" x ")
                qty_str = parts[0].strip().split(" ")[0]
                if qty_str.replace(',','').isdigit():
                    qty = int(qty_str.replace(',',''))
            except Exception:
                pass

        retail_price = 0.0
        if monthly_cost:
            try:
                retail_price = float(monthly_cost) / (qty if qty > 0 else 1)
            except:
                pass

        item = {
            "group_path": group_path,
            "service_name": svc_type or cat,
            "applied": custom_name or svc_type or cat,
            "spec": desc or "",
            "qty": qty,
            "unit": unit,
            "region": region,
            "retailPrice": retail_price,
            "billing_option": "Pay-as-you-go"
        }
        inventory.append(item)
    return inventory

uploaded_file = st.file_uploader("Azure Estimate 파일 업로드 (.xlsx)", type=['xlsx'])

if uploaded_file is not None:
    st.success("파일 업로드 완료! 변환을 시작합니다.")
    
    with st.spinner("엑셀 데이터를 분석하고 변환하는 중..."):
        try:
            # 1. 파일 파싱해서 inventory_items 추출
            inventory = parse_azure_export(uploaded_file)
            
            # 2. sample_data.json 덮어쓰기
            data = {
                "project_info": {
                    "customer_name": "고객사 (자동생성)",
                    "project_name": "Cloud 인프라 환경 구축",
                    "fx_rate": 1430,
                    "vat_included": False
                },
                "inventory_items": inventory
            }
            with open("sample_data.json", "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
            # 3. 엑셀 생성 스크립트 강제 호출
            output_filename = "Converted_Estimate_Result.xlsx"
            if os.path.exists(output_filename):
                os.remove(output_filename)
                
            # generate_quote의 메인 함수 실행
            generate_quote.run_generation(output_filename)
            
            st.success("변환이 성공적으로 완료되었습니다!")
            
            # 4. 다운로드 버튼 제공
            with open(output_filename, "rb") as f:
                st.download_button(
                    label="📥 변환된 견적서 다운로드",
                    data=f,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
        except Exception as e:
            st.error(f"변환 중 오류가 발생했습니다: {e}")
