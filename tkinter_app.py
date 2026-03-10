import tkinter as tk
from tkinter import filedialog, messagebox
import os
import json
from openpyxl import load_workbook
import generate_quote

def parse_azure_export(file_path):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.worksheets[0]
    
    inventory = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row[0]: 
            if len(row) > 3 and row[3] == 'Total':
                break
            continue
            
        cat, svc_type, custom_name, region, desc, monthly_cost, upfront = row[:7]
        
        main_group = "기타자원"
        sub_group = cat
        
        if cat in ["컴퓨팅", "Virtual Machines", "데이터베이스", "Databases", "Containers"]: 
            main_group = "Server 및 DBMS"
            sub_group = cat
        elif cat in ["Storage", "저장소"]: 
            main_group = "Storage"
            sub_group = "Storage Accounts"
        elif cat in ["Networking", "네트워크"]: 
            main_group = "Network"
            sub_group = "Network"
        elif cat in ["Management and Governance"]:
            main_group = "모니터링"
            sub_group = "Azure Monitor"
        elif (cat and "Backup" in cat) or (svc_type and "Backup" in svc_type):
            main_group = "기타자원"
            sub_group = "Backup"
            
        group_path = f"{main_group} > {sub_group}"
        
        qty = 1
        unit = "ea"
        if isinstance(desc, str) and " x " in desc:
            try:
                parts = desc.split(" x ")
                qty_str = parts[0].strip().split(" ")[0]
                if qty_str.replace(',','').isdigit():
                    qty = int(qty_str.replace(',',''))
            except Exception:
                pass

        retail_price = 0.0
        if monthly_cost:
            try:
                retail_price = float(monthly_cost) / (qty if qty > 0 else 1)
            except:
                pass

        item = {
            "group_path": group_path,
            "service_name": svc_type or cat,
            "applied": custom_name or svc_type or cat,
            "spec": desc or "",
            "qty": qty,
            "unit": unit,
            "region": region,
            "retailPrice": retail_price,
            "billing_option": "Pay-as-you-go"
        }
        inventory.append(item)
    return inventory

def run_conversion():
    file_path = filedialog.askopenfilename(
        title="Azure 계산기 엑셀 파일 선택",
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    if not file_path:
        return
        
    try:
        status_var.set("엑셀 데이터를 분석하고 변환하는 중...")
        root.update()
        
        inventory = parse_azure_export(file_path)
        
        data = {
            "project_info": {
                "customer_name": "고객사 (자동생성)",
                "project_name": "Cloud 인프라 환경 구축",
                "fx_rate": 1430,
                "vat_included": False
            },
            "inventory_items": inventory
        }
        with open("sample_data.json", "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            
        output_filename = "가상의_고객_주식회사_견적서_결과.xlsx"
        if os.path.exists(output_filename):
            os.remove(output_filename)
            
        generate_quote.run_generation(output_filename)
        
        status_var.set(f"변환 완료!\n생성된 파일: {output_filename}")
        messagebox.showinfo("성공", f"의도한 템플릿으로 엑셀 변환이 완벽하게 끝났습니다!\n\n저장 경로:\n{os.path.abspath(output_filename)}")
        
    except Exception as e:
        status_var.set("변환 실패.")
        messagebox.showerror("오류 발생", f"파일을 변환하는 도중 문제가 생겼습니다:\n{e}")

# UI Setup
root = tk.Tk()
root.title("Azure 견적서 자동 변환기")
root.geometry("450x250")
root.configure(padx=20, pady=20)

title_lbl = tk.Label(root, text="☁️ Azure Estimate to SSG Quote", font=("맑은 고딕", 14, "bold"))
title_lbl.pack(pady=10)

desc_lbl = tk.Label(root, text="Azure 계산기에서 내보낸 엑셀(ExportedEstimate.xlsx)을\n선택하시면 자동으로 단일 탭 견적서로 변환합니다.", font=("맑은 고딕", 10), justify="center")
desc_lbl.pack(pady=10)

upload_btn = tk.Button(root, text="엑셀 파일 선택 및 변환 시작", command=run_conversion, font=("맑은 고딕", 11, "bold"), bg="#0066cc", fg="white", width=25, height=2)
upload_btn.pack(pady=10)

status_var = tk.StringVar()
status_var.set("대기 중...")
status_lbl = tk.Label(root, textvariable=status_var, font=("맑은 고딕", 9), fg="gray")
status_lbl.pack(side="bottom", pady=5)

root.mainloop()
