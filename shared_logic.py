from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, Color
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_interval
import os
import urllib.parse
import urllib.request
import json
import re

def fetch_azure_price(region, service_name, meter_name, sku_name, product_name):
    filters = [
        f"armRegionName eq '{region}'",
        f"serviceName eq '{service_name}'",
        f"priceType eq 'Consumption'"
    ]
    if meter_name: filters.append(f"meterName eq '{meter_name}'")
    if sku_name: filters.append(f"skuName eq '{sku_name}'")
    if product_name: filters.append(f"productName eq '{product_name}'")
    
    filter_str = ' and '.join(filters)
    url = f"https://prices.azure.com/api/retail/prices?$filter={urllib.parse.quote(filter_str)}"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode())
            items = data.get('Items', [])
            if items:
                return float(items[0].get('retailPrice', 0))
    except Exception as e:
        print(f"Error fetching price: {e}")
    return 0.0

# ── Styles & Utils ───────────────────────────────────────────────────────────

def S(style=None):
    return Side(style=style)

def brd(l=None, r=None, t=None, b=None):
    return Border(left=S(l), right=S(r), top=S(t), bottom=S(b))

def rgb_fill(hex6):
    return PatternFill('solid', fgColor=f'FF{hex6}')

def theme_fill(n=0, tint=0.0):
    c = Color(theme=n, tint=tint, type='theme')
    pf = PatternFill(patternType='solid')
    pf.fgColor = c
    return pf

def tf0():
    return theme_fill(0, 0.0)

def tf0d():
    return theme_fill(0, -0.1499984740745262)

def font(bold=False, size=10, name='맑은 고딕'):
    return Font(name=name, bold=bold, size=size)

def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def sc(ws, coord, val=None, fnt=None, al=None, border=None, fill=None, nf=None):
    c = ws[coord]
    if val is not None:
        c.value = val
    if fnt:
        c.font = fnt
    if al:
        c.alignment = al
    if border is not None:
        c.border = border
    if fill is not None:
        c.fill = fill
    if nf:
        c.number_format = nf
    return c

def apply_border_to_range(ws, cell_range, border):
    if ":" not in cell_range:
        ws[cell_range].border = border
        return
    match = re.search(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", cell_range)
    if not match: return
    start_col, start_row, end_col, end_row = match.groups()
    for row in range(int(start_row), int(end_row) + 1):
        for col in get_column_interval(start_col, end_col):
            ws[f"{col}{row}"].border = border

WON_FMT  = '_-"₩"* #,##0_-;\\-"₩"* #,##0_-;_-"₩"* "-"_-;_-@_-'
WON_FMT2 = '"₩"#,##0_);[Red]\\("₩"#,##0\\)'
USD_FMT  = '\\$#,##0.00_);[Red]\\(\\$#,##0.00\\)'
QTY_FMT  = '0_);[Red]\\(0\\)'

T4 = brd('thin','thin','thin','thin')
H4 = brd('hair','hair','hair','hair')
M4 = brd('medium','medium','medium','medium')

# ── Excel Porting ────────────────────────────────────────────────────────────

def port_styles_from_original(output_path, orig_path):
    import zipfile, os
    from xml.etree import ElementTree as ET
    NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    with zipfile.ZipFile(orig_path) as z:
        orig_styles = z.read('xl/styles.xml')
        def extract_maps(xml_bytes):
            s_map, v_map = {}, {}
            root = ET.fromstring(xml_bytes)
            for row_elem in root.findall(f'.//{NS}row'):
                for cell in row_elem.findall(f'{NS}c'):
                    ref = cell.get('r')
                    s = cell.get('s')
                    if s: s_map[ref] = s
                    v_elem = cell.find(f'{NS}v')
                    if cell.find(f'{NS}f') is not None and v_elem is not None and v_elem.text:
                        v_map[ref] = (v_elem.text, cell.get('t', ''))
            return s_map, v_map
        orig_s1, orig_v1 = extract_maps(z.read('xl/worksheets/sheet1.xml'))

    with zipfile.ZipFile(output_path) as z:
        files = {name: z.read(name) for name in z.namelist()}
    files['xl/styles.xml'] = orig_styles

    def patch_sheet(sheet_xml_bytes, s_map, v_map):
        root = ET.fromstring(sheet_xml_bytes)
        for row_elem in root.findall(f'.//{NS}row'):
            for cell in row_elem.findall(f'{NS}c'):
                ref = cell.get('r')
                if ref in s_map: cell.set('s', s_map[ref])
                if ref in v_map:
                    v_text, t_attr = v_map[ref]
                    v_elem = cell.find(f'{NS}v')
                    if v_elem is None: v_elem = ET.SubElement(cell, f'{NS}v')
                    v_elem.text = v_text
                    if t_attr: cell.set('t', t_attr)
        return ET.tostring(root, xml_declaration=True, encoding='UTF-8')

    if 'xl/worksheets/sheet1.xml' in files:
        files['xl/worksheets/sheet1.xml'] = patch_sheet(files['xl/worksheets/sheet1.xml'], orig_s1, orig_v1)

    tmp = output_path + '.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items(): zout.writestr(name, data)
    os.replace(tmp, output_path)

# ── Sheet Builders ───────────────────────────────────────────────────────────

def build_sheet1(wb, srt, ort, mrt, project_info=None):
    ws = wb.active # 첫 시트
    ws.title = '매출견적서'
    cust_name = project_info.get('customer_name', 'SK렌터카') if project_info else 'SK렌터카'
    proj_name = project_info.get('project_name', 'Azure Resource') if project_info else 'Azure Resource'

    for col, w in [('A',1.375),('B',11.625),('C',32.625),('D',47.5),('E',6.125),('F',15.625),('G',13.0),('H',53.375)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:H2')
    sc(ws,'B2','見   積   書', fnt=Font(name='맑은 고딕',bold=True,size=26), al=aln('center','center'))

    for r, txt in [(6,f'회   사   명 : {cust_name}'),(7,'견 적 번 호 : 가견적'),(8,'견 적 날 짜 : 2026년  3월  3일'),
                   (9,'견적 유효 기간 : 견적일로부터 30일'),(10,'납         기 : 별도 협의')]:
        sc(ws,f'B{r}',txt,fnt=font(),al=aln('left','center'))

    ws.merge_cells('E6:E11')
    sc(ws,'E6','공\n\n급\n\n자',fnt=font(),al=aln('center','center',wrap=True), border=H4, fill=tf0d())
    apply_border_to_range(ws, 'E6:E11', H4)

    for r in range(6,12): ws.merge_cells(f'G{r}:H{r}')
    for coord, txt in [('F6','등록번호'),('F7','상       호'),('F8','대표자명'),('F9','주       소'),('F10','업       태'),('F11','종       목')]:
        sc(ws,coord,txt,fnt=font(),al=aln('left','center'),border=H4,fill=tf0d())
    for r, txt in enumerate(['  783-85-00169', '  SK㈜', '  김완종 (직인생략)', '  성남시 분당구 강남대로 343번길 9', '  서비스(사업관련)', '  기타정보처리및컴퓨터 운용관련']):
        sc(ws,f'G{6+r}',txt,fnt=font(),al=aln('left','center'),border=brd('hair','hair','hair','hair'),fill=tf0d())
        apply_border_to_range(ws, f'G{6+r}:H{6+r}', brd('hair','hair','hair','hair'))

    sc(ws,'B14','제   목 : ',fnt=font(),al=aln('left','center'))
    sc(ws,'C14',proj_name,fnt=font(),al=aln('left','center'))
    sc(ws,'H14','(금액단위 : 원/VAT 별도)',fnt=font(),al=aln('right','center'))

    hf = rgb_fill('F99107')
    ws.merge_cells('C15:D15')
    for coord, txt in [('B15','No.'),('E15','수량\n(개월)'),('F15','월 공급금액'),('G15','총 공급금액\n(1년 간)'),('H15','비고')]:
        sc(ws,coord,txt,fnt=font(bold=True),al=aln('center','center',wrap=True),border=T4,fill=hf)
    sc(ws,'C15','품목 및 사양',fnt=font(bold=True),al=aln('center','center'),border=T4,fill=hf)
    apply_border_to_range(ws, 'C15:D15', T4)

    ws.merge_cells('B16:B19')
    sc(ws,'B16','Azure\nResource',fnt=font(bold=True), al=aln('center','center',wrap=True),border=T4,fill=tf0d())
    apply_border_to_range(ws, 'B16:B19', T4)

    rows_az = [
        (16,'Server','Cloud 사용료',12,f"='Azure Resource'!K{srt}",'=F16*E16','-Azure Resource List Price의 10% 할인 적용\n(Pay-as-you-go 상품에 한하며, RI상품은 5%할인 적용됨)'),
        (17,'기타 자원(Storage/DB 등)','Cloud 사용료',12,f"='Azure Resource'!K{ort}",'=F17*E17',None),
        (18,'모니터링(Datadog)','Cloud 사용료',12,f"='Azure Resource'!K{mrt}",'=F18*E18',None),
        (19,'Azure Support Plan','SK Enterprise Support for Azure - SKN 계열 통합 과금',12,'=SUM(F16:F17)*0.095','=F19*E19',None),
    ]
    for row,c,d,e,f,g,h in rows_az:
        for col, val in zip(['C','D','E','F','G','H'], [c,d,e,f,g,h]):
            sc(ws,f'{col}{row}',val,fnt=font(),al=aln('center' if col!='H' else 'left','center',wrap=True),border=T4)
    ws.merge_cells('H16:H17')
    apply_border_to_range(ws, 'H16:H17', T4)

    ws.merge_cells('B20:E20')
    sc(ws,'B20','Azure Resource 합계',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=tf0d())
    apply_border_to_range(ws, 'B20:E20', T4)
    sc(ws,'F20','=ROUNDDOWN(SUM(F16:F19),-3)',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=tf0d(),nf=WON_FMT)
    sc(ws,'G20','=ROUNDDOWN(SUM(G16:G19),-3)',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=tf0d(),nf=WON_FMT)
    sc(ws,'H20',border=T4,fill=tf0d())

    ws.merge_cells('B21:B23')
    sc(ws,'B21','MSP\n운영비',fnt=font(bold=True), al=aln('center','center',wrap=True),border=T4,fill=tf0d())
    apply_border_to_range(ws, 'B21:B23', T4)
    for r in range(21, 24):
        for col in ['C','D','E','F','G','H']: sc(ws,f'{col}{r}',border=T4)

    ws.merge_cells('B24:E24')
    sc(ws,'B24','MSP 운영비 합계',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=tf0d())
    apply_border_to_range(ws, 'B24:E24', T4)
    sc(ws,'F24','=ROUNDDOWN(SUM(F21:F23),-3)',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=tf0d(),nf=WON_FMT)
    sc(ws,'G24','=ROUNDDOWN(SUM(G21:G23),-3)',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=tf0d(),nf=WON_FMT)
    sc(ws,'H24',border=T4,fill=tf0d())

    ws.merge_cells('B25:H25')
    sc(ws,'B25','이하 여백',fnt=font(),al=aln('center','center'), border=T4,fill=tf0())
    apply_border_to_range(ws, 'B25:H25', T4)

    ws.merge_cells('B26:E26')
    yf = rgb_fill('FFFF99')
    sc(ws,'B26','사용료 합계(부가세 별도)',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=yf)
    apply_border_to_range(ws, 'B26:E26', T4)
    sc(ws,'F26','=SUM(F20,F24)',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=yf,nf=WON_FMT)
    sc(ws,'G26','=SUM(G20,G24)',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=yf,nf=WON_FMT)
    sc(ws,'H26',border=T4,fill=yf)

    ws.merge_cells('B27:E27')
    of = rgb_fill('FFCC99')
    sc(ws,'B27','최종 사용료 - 최종 계약금(VAT 별도)',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=of)
    apply_border_to_range(ws, 'B27:E27', T4)
    sc(ws,'F27','=F26',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=of,nf=WON_FMT)
    sc(ws,'G27','=G26',fnt=font(bold=True), al=aln('center','center'),border=T4,fill=of,nf=WON_FMT)
    sc(ws,'H27',border=T4,fill=of)

    sc(ws,'B28','※ Remark',fnt=font(bold=True))
    sc(ws,'B29','- 상기 비용은 실 사용량 및 환율 변동에 따라 금액이 변동될 수 있음.',fnt=font())

def build_sheet2(wb, inventory, project_info=None):
    fx_rate = project_info.get('fx_rate', 1430) if project_info else 1430
    proj_name = project_info.get('project_name', 'Azure Resource') if project_info else 'Azure Resource'
    
    ws = wb.create_sheet('Azure Resource')
    for col, w in [('A',1.625),('B',11.625),('C',32.875),('D',24.125),('E',26.375),('F',44.25),('G',6.25),('H',5.875),('I',12.875),('J',13.0),('K',15.125),('L',53.375)]:
        ws.column_dimensions[col].width = w

    sc(ws,'B2',f'Cloud 세부내역 : {proj_name}', fnt=font(bold=True,size=12),al=aln('left','center'))
    sc(ws,'K3','환율',fnt=font(bold=True),al=aln('center','center'),border=T4)
    sc(ws,'K4',fx_rate,fnt=font(),al=aln('center','center'),border=T4)

    hf = rgb_fill('F99107')
    for rng in ['B5:B6','C5:F5','G5:G6','H5:H6','I5:I6','J5:J6','K5:K6','L5:L6']:
        ws.merge_cells(rng)
    ws.merge_cells('D6:F6')

    # 5행 Header (수동 테두리 지정으로 빈 테두리 방지)
    sc(ws,'B5','유형',fnt=font(bold=True),al=aln('center','center',wrap=True),border=brd('medium','thin','medium',None),fill=hf)
    sc(ws,'C5','품목 및 사양',fnt=font(bold=True),al=aln('center','center'),border=brd('thin',None,'medium',None),fill=hf)
    for col in ['D','E','F']: sc(ws,f'{col}5',border=brd(None,None,'medium',None))
    sc(ws,'G5','수량',fnt=font(bold=True),al=aln('center','center',wrap=True),border=brd('thin',None,'medium',None),fill=hf)
    for coord, txt in [('H5','단위'),('I5','Price\n($/Month)'),('J5','Price\n(\\/Month)'),('K5','C&C 공급금액\n(\\/Month)')]:
        sc(ws,coord,txt,fnt=font(bold=True),al=aln('center','center',wrap=True),border=brd('thin','thin','medium',None),fill=hf)
    sc(ws,'L5','비고',fnt=font(bold=True),al=aln('center','center',wrap=True),border=brd(None,'medium','medium',None),fill=theme_fill(9))

    # 6행 Header
    sc(ws,'B6',border=brd('medium','thin',None,'medium'))
    sc(ws,'C6','서비스 명',fnt=font(bold=True),al=aln('center','center'),border=brd('thin',None,'thin','medium'),fill=hf)
    sc(ws,'D6','상세 Spec',fnt=font(bold=True),al=aln('center','center'),border=brd(None,None,'thin','medium'),fill=hf)
    sc(ws,'E6',border=brd(None,None,'thin','medium'))
    sc(ws,'F6',border=brd(None,'thin','thin','medium'))
    sc(ws,'G6',border=brd('thin',None,None,'medium'))
    for col in ['H','I','J','K']:
        sc(ws,f'{col}6',border=brd('thin','thin',None,'medium'))
    sc(ws,'L6',border=brd(None,'medium',None,'medium'))

    hierarchy = {}
    for item in inventory:
        parts = [p.strip() for p in item.get('group_path', '기타 > 기타').split('>')]
        main_cat = parts[0]
        sub_cat = parts[1] if len(parts) > 1 else '기타'
        if main_cat not in hierarchy: hierarchy[main_cat] = {}
        if sub_cat not in hierarchy[main_cat]: hierarchy[main_cat][sub_cat] = []
        hierarchy[main_cat][sub_cat].append(item)

    current_row = 7
    subtotals = []
    
    for main_cat, sub_dict in hierarchy.items():
        main_start = current_row
        for sub_cat, items_list in sub_dict.items():
            sub_start = current_row
            sub_end = current_row + len(items_list) - 1
            ws.merge_cells(f'C{sub_start}:C{sub_end}')
            
            c_bottom = 'medium' if sub_start == sub_end else None
            c_top = 'medium' if sub_start == main_start else 'thin'
            sc(ws, f'C{sub_start}', sub_cat, fnt=font(bold=True), al=aln('center','center'), border=brd('thin','thin',c_top,c_bottom))
            for r in range(sub_start+1, sub_end+1):
                sc(ws, f'C{r}', border=brd('thin','thin',None,'medium' if r==sub_end else None))

            for idx, item in enumerate(items_list):
                r = current_row
                is_first = (sub_start == main_start and idx == 0)
                is_last = (r == sub_end)
                
                t_style = 'medium' if is_first else 'thin'
                b_style = 'medium' if is_last else 'thin'
                
                price_usd = item.get('retailPrice', 0)
                d_val = item.get('service_name')
                e_val = item.get('applied')
                f_val = f"{item.get('spec')}\n{item.get('note', '')}"
                
                sc(ws,f'D{r}',d_val,fnt=font(),al=aln('center','center'), border=brd('thin','thin',t_style,b_style))
                sc(ws,f'E{r}',e_val,fnt=font(),al=aln('center','center'), border=brd('thin','hair',t_style,b_style))
                sc(ws,f'F{r}',f_val,fnt=font(),al=aln('left','center',wrap=True), border=brd('hair','thin',t_style,b_style))
                
                sc(ws,f'G{r}',item.get('qty', 1),fnt=font(),al=aln('center','center'), border=brd('thin','thin',t_style,b_style))
                sc(ws,f'H{r}',item.get('unit', 'ea'),fnt=font(),al=aln('center','center'), border=brd('thin','thin',t_style,b_style))
                sc(ws,f'I{r}',price_usd,fnt=font(),al=aln('center','center'), border=brd('thin','thin',t_style,b_style), nf=USD_FMT)
                sc(ws,f'J{r}',f'=I{r}*$K$4*G{r}', fnt=font(),al=aln('center','center'), border=brd('thin','thin',t_style,b_style), nf=WON_FMT2)
                sc(ws,f'K{r}',f'=J{r}*(1-10%)', fnt=font(),al=aln('center','center'), border=brd('thin','thin',t_style,b_style), nf=WON_FMT2)
                sc(ws,f'L{r}',border=brd(None,'medium',t_style,b_style))
                
                current_row += 1

        main_end = current_row - 1
        ws.merge_cells(f'B{main_start}:B{main_end}')
        sc(ws,f'B{main_start}',main_cat,fnt=font(bold=True), al=aln('center','center',wrap=True), border=brd('medium','thin','medium',None if main_start != main_end else 'medium'))
        for r in range(main_start+1, main_end+1):
            sc(ws,f'B{r}', border=brd('medium','thin',None,'medium' if r==main_end else None))

        subtotal_row = current_row
        ws.merge_cells(f'B{current_row}:J{current_row}')
        sc(ws,f'B{current_row}',f'{main_cat} 합계 :',fnt=font(bold=True), al=aln('right','center'), border=brd('medium',None,'medium','medium'),fill=theme_fill(5, 0.6))
        for col in ['C','D','E','F','G','H','I','J']:
            b_right = 'medium' if col == 'J' else None
            sc(ws,f'{col}{current_row}', border=brd(None,b_right,'medium','medium'), fill=theme_fill(5, 0.6))
        
        sc(ws,f'K{current_row}',f'=SUM(K{main_start}:K{main_end})',fnt=font(bold=True), al=aln('center','center'), border=brd('medium','thin','medium','medium'),fill=theme_fill(5, 0.6),nf=WON_FMT2)
        sc(ws,f'L{current_row}',border=brd('thin','medium','medium','medium'),fill=theme_fill(5, 0.6))
        
        subtotals.append((main_cat, subtotal_row))
        current_row += 1

    # Total Row
    ws.merge_cells(f'B{current_row}:J{current_row}')
    sc(ws,f'B{current_row}','총 합계(백신제외) :',fnt=font(bold=True), al=aln('right','center'), border=brd('medium',None,'medium','medium'),fill=rgb_fill('FFC000'))
    for col in ['C','D','E','F','G','H','I','J']:
        b_right = 'medium' if col == 'J' else None
        sc(ws,f'{col}{current_row}', border=brd(None,b_right,'medium','medium'), fill=rgb_fill('FFC000'))
    
    sub_terms = ",".join([f"K{r}" for _, r in subtotals])
    sc(ws,f'K{current_row}',f"=SUM({sub_terms})" if sub_terms else "=0",fnt=font(bold=True), al=aln('center','center'), border=brd('medium','thin','medium','medium'),fill=rgb_fill('FFC000'),nf=WON_FMT2)
    sc(ws,f'L{current_row}',border=brd('thin','medium','medium','medium'),fill=rgb_fill('FFC000'))

    current_row += 2
    for r in range(current_row, current_row+3): ws.merge_cells(f'B{r}:L{r}')
    sc(ws,f'B{current_row}','※ Remark',fnt=font(bold=True), al=aln('left','center'))
    sc(ws,f'B{current_row+1}',f'- 상기 금액은 환율 {fx_rate:.2f}원 기준임.',fnt=font(), al=aln('left','center'))
    ws.row_dimensions[current_row+2].height = 45

    srt, ort, mrt = 20, 20, 20
    for cat, r in subtotals:
        if 'Server' in cat: srt = r
        elif '기타' in cat or 'Storage' in cat or 'Database' in cat: ort = r
    return srt, ort, mrt
