import os
import json
from openpyxl import load_workbook
import generate_quote

def parse_azure_export(file_path):
    # logic copied from app.py to ensure parity with existing development
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.worksheets[0]
    
    inventory = []
    # 데이터는 보통 4행부터 시작 (1, 2행 제목, 3행 헤더)
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row[0]: # Service category null 이면 스킵
            if len(row) > 3 and row[3] == 'Total':
                break
            continue
            
        cat, svc_type, custom_name, region, desc, monthly_cost, upfront = row[:7]
        
        main_group = "기타자원"
        if cat in ["컴퓨팅", "Virtual Machines", "데이터베이스", "Databases", "Containers"]: 
            main_group = "Server 및 DBMS"
        elif cat in ["Storage", "저장소"]: 
            main_group = "Storage"
        elif cat in ["Networking", "네트워크"]: 
            main_group = "Network"
        elif cat in ["Management and Governance"]:
            main_group = "모니터링"
            
        group_path = f"{main_group} > {cat}"
        
        qty = 1
        unit = "ea"
        if isinstance(desc, str) and " x " in desc:
            try:
                parts = desc.split(" x ")
                qty_str = parts[0].strip().split(" ")[0]
                if qty_str.replace(',','').isdigit():
                    qty = int(qty_str.replace(',',''))
            except Exception:
                pass

        retail_price = 0.0
        if monthly_cost:
            try:
                retail_price = float(monthly_cost) / (qty if qty > 0 else 1)
            except:
                pass

        item = {
            "group_path": group_path,
            "service_name": svc_type or cat,
            "applied": custom_name or svc_type or cat,
            "spec": desc or "",
            "qty": qty,
            "unit": unit,
            "region": region,
            "retailPrice": retail_price,
            "billing_option": "Pay-as-you-go"
        }
        inventory.append(item)
    return inventory

def run():
    input_file = os.path.join("inputs", "ExportedEstimate_1.xlsx")
    output_file = "가상의_고객_주식회사_견적서_결과.xlsx"
    
    print(f"Parsing {input_file}...")
    inventory = parse_azure_export(input_file)
    
    # Update sample_data.json as per app.py logic
    data = {
        "project_info": {
            "customer_name": "고객사 (자동생성)",
            "project_name": "Cloud 인프라 환경 구축",
            "fx_rate": 1430,
            "vat_included": False
        },
        "inventory_items": inventory
    }
    with open("sample_data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"Generating quote: {output_file}...")
    generate_quote.run_generation(output_file)
    print("Done!")

if __name__ == "__main__":
    run()
